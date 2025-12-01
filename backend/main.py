"""
VendHub Database - FastAPI Backend
Основной файл приложения
"""
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File as FastAPIFile, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import List, Optional
from datetime import datetime, timedelta, date
from pydantic import BaseModel
from loguru import logger
import os
import io

from database import get_db, init_db
from models import User, File, Record
from auth import (
    authenticate_user, create_user, create_access_token,
    get_current_user, UserCreate, UserResponse, Token
)
from utils.excel_parser import (
    parse_excel_file, extract_date_from_row,
    get_period_from_date, validate_excel_structure
)

# Инициализация приложения
app = FastAPI(
    title="VendHub Database API",
    description="API для управления базой данных VendHub",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=os.getenv("ALLOWED_ORIGINS", "*").split(","),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Логирование
logger.add("logs/vendhub_{time}.log", rotation="1 day", retention="7 days")


# Pydantic схемы
class FileResponse(BaseModel):
    id: int
    filename: str
    row_count: int
    uploaded_at: datetime
    headers: Optional[List[str]]

    class Config:
        from_attributes = True


class RecordResponse(BaseModel):
    id: int
    file_id: int
    data: dict
    date_field: Optional[date]
    period: Optional[str]

    class Config:
        from_attributes = True


class PeriodStats(BaseModel):
    period: str
    count: int
    label: str


class DatabaseStats(BaseModel):
    total_files: int
    total_records: int
    unique_records: int
    periods: List[PeriodStats]


# ========================================
# События приложения
# ========================================

@app.on_event("startup")
async def startup_event():
    """Инициализация при запуске"""
    logger.info("Starting VendHub Database API...")
    init_db()
    logger.info("Database initialized successfully")


@app.on_event("shutdown")
async def shutdown_event():
    """Очистка при выключении"""
    logger.info("Shutting down VendHub Database API...")


# ========================================
# Аутентификация
# ========================================

@app.post("/api/auth/register", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def register(user_data: UserCreate, db: Session = Depends(get_db)):
    """Регистрация нового пользователя"""
    try:
        user = create_user(db, user_data)
        logger.info(f"New user registered: {user.username}")
        return user
    except HTTPException as e:
        raise e
    except Exception as e:
        logger.error(f"Registration error: {str(e)}")
        raise HTTPException(status_code=500, detail="Registration failed")


@app.post("/api/auth/login", response_model=Token)
async def login(
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    """Вход пользователя"""
    user = authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    access_token = create_access_token(data={"sub": user.username})
    logger.info(f"User logged in: {user.username}")

    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/api/auth/me", response_model=UserResponse)
async def get_me(current_user: User = Depends(get_current_user)):
    """Получить информацию о текущем пользователе"""
    return current_user


# ========================================
# Файлы
# ========================================

@app.post("/api/files/upload", response_model=FileResponse, status_code=status.HTTP_201_CREATED)
async def upload_file(
    file: UploadFile = FastAPIFile(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Загрузка Excel файла"""

    # Проверка типа файла
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are allowed"
        )

    try:
        # Чтение файла
        file_content = await file.read()

        # Парсинг
        headers, rows = parse_excel_file(file_content)

        if len(rows) == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File contains no data rows"
            )

        # Создание записи файла
        db_file = File(
            user_id=current_user.id,
            filename=file.filename,
            row_count=len(rows),
            headers=headers
        )
        db.add(db_file)
        db.flush()  # Получаем ID файла

        # Обработка строк
        records_added = 0
        for row in rows:
            # Извлекаем дату
            date_obj = extract_date_from_row(row)
            period = get_period_from_date(date_obj) if date_obj else None

            # Создаем словарь данных
            row_data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            # Проверяем уникальность (file_id + data)
            existing = db.query(Record).filter(
                Record.file_id == db_file.id,
                Record.data == row_data
            ).first()

            if not existing:
                record = Record(
                    file_id=db_file.id,
                    data=row_data,
                    date_field=date_obj,
                    period=period
                )
                db.add(record)
                records_added += 1

        db.commit()
        db.refresh(db_file)

        logger.info(f"File uploaded: {file.filename}, {records_added} records added by user {current_user.username}")

        return db_file

    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        logger.error(f"File upload error: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail="File processing failed")


@app.get("/api/files", response_model=List[FileResponse])
async def get_files(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить список файлов пользователя"""
    files = db.query(File).filter(File.user_id == current_user.id).order_by(File.uploaded_at.desc()).all()
    return files


@app.get("/api/files/{file_id}", response_model=FileResponse)
async def get_file(
    file_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить информацию о файле"""
    file = db.query(File).filter(
        File.id == file_id,
        File.user_id == current_user.id
    ).first()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    return file


@app.delete("/api/files/{file_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_file(
    file_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Удалить файл и его данные"""
    file = db.query(File).filter(
        File.id == file_id,
        File.user_id == current_user.id
    ).first()

    if not file:
        raise HTTPException(status_code=404, detail="File not found")

    db.delete(file)
    db.commit()

    logger.info(f"File deleted: {file.filename} by user {current_user.username}")
    return None


# ========================================
# Записи (база данных)
# ========================================

@app.get("/api/records", response_model=dict)
async def get_records(
    search: Optional[str] = Query(None, description="Поиск по всем полям"),
    period: Optional[str] = Query(None, description="Фильтр по периоду (YYYY-MM)"),
    date_from: Optional[date] = Query(None, description="Дата с"),
    date_to: Optional[date] = Query(None, description="Дата по"),
    page: int = Query(1, ge=1, description="Номер страницы"),
    size: int = Query(50, ge=1, le=500, description="Размер страницы"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить записи с фильтрацией и пагинацией"""

    # Базовый запрос
    query = db.query(Record).join(File).filter(File.user_id == current_user.id)

    # Фильтрация
    if search:
        # Поиск по JSONB данным
        query = query.filter(Record.data.astext.ilike(f"%{search}%"))

    if period:
        query = query.filter(Record.period == period)

    if date_from:
        query = query.filter(Record.date_field >= date_from)

    if date_to:
        query = query.filter(Record.date_field <= date_to)

    # Подсчет всего
    total = query.count()

    # Пагинация
    offset = (page - 1) * size
    records = query.order_by(Record.date_field.desc().nullslast()).offset(offset).limit(size).all()

    # Формирование ответа
    return {
        "total": total,
        "page": page,
        "size": size,
        "pages": (total + size - 1) // size,
        "data": [
            {
                "id": r.id,
                "file_id": r.file_id,
                "data": r.data,
                "date_field": r.date_field,
                "period": r.period
            }
            for r in records
        ]
    }


@app.get("/api/records/stats", response_model=DatabaseStats)
async def get_database_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить статистику базы данных"""

    # Статистика файлов
    total_files = db.query(func.count(File.id)).filter(File.user_id == current_user.id).scalar()

    # Статистика записей
    total_records = db.query(func.count(Record.id)).join(File).filter(File.user_id == current_user.id).scalar()

    # Уникальные записи (по data)
    unique_records = db.query(func.count(func.distinct(Record.data))).join(File).filter(
        File.user_id == current_user.id
    ).scalar()

    # Статистика по периодам
    period_stats = db.query(
        Record.period,
        func.count(Record.id).label('count')
    ).join(File).filter(
        File.user_id == current_user.id,
        Record.period.isnot(None)
    ).group_by(Record.period).order_by(Record.period.desc()).all()

    # Форматирование периодов
    month_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']

    periods = []
    for period, count in period_stats:
        year, month = period.split('-')
        label = f"{month_names[int(month) - 1]} {year}"
        periods.append(PeriodStats(period=period, count=count, label=label))

    return DatabaseStats(
        total_files=total_files or 0,
        total_records=total_records or 0,
        unique_records=unique_records or 0,
        periods=periods
    )


@app.get("/api/records/export")
async def export_records(
    search: Optional[str] = Query(None),
    period: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт записей в Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    # Получаем данные с фильтрами (без пагинации)
    query = db.query(Record).join(File).filter(File.user_id == current_user.id)

    if search:
        query = query.filter(Record.data.astext.ilike(f"%{search}%"))
    if period:
        query = query.filter(Record.period == period)
    if date_from:
        query = query.filter(Record.date_field >= date_from)
    if date_to:
        query = query.filter(Record.date_field <= date_to)

    records = query.order_by(Record.date_field.desc().nullslast()).all()

    if not records:
        raise HTTPException(status_code=404, detail="No data to export")

    # Создаем Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VendHub Database"

    # Заголовки (из первой записи)
    headers = list(records[0].data.keys())
    ws.append(['#'] + headers)

    # Стиль заголовков
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Данные
    for i, record in enumerate(records, 1):
        row_data = [i] + [record.data.get(h, '') for h in headers]
        ws.append(row_data)

    # Сохранение в буфер
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"VendHub_Export_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    logger.info(f"Exported {len(records)} records for user {current_user.username}")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ========================================
# Health Check
# ========================================

@app.get("/health")
async def health_check():
    """Проверка состояния API"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0"
    }


@app.get("/api/admin/reset-db")
async def reset_database(secret: str = Query(...), db: Session = Depends(get_db)):
    """Одноразовая очистка БД и создание админа (удалить после использования)"""
    if secret != "vendhub-reset-2024":
        raise HTTPException(status_code=403, detail="Invalid secret")

    from passlib.context import CryptContext

    try:
        # Удаляем все записи, файлы и пользователей
        db.query(Record).delete()
        db.query(File).delete()
        db.query(User).delete()
        db.commit()

        # Создаем админа
        pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
        admin_user = User(
            username="jamshiddin",
            email="admin@vendhub.com",
            password_hash=pwd_context.hash("311941990")
        )
        db.add(admin_user)
        db.commit()

        logger.info("Database reset completed. Admin user created.")
        return {"status": "success", "message": "Database cleared. Admin user 'jamshiddin' created."}
    except Exception as e:
        db.rollback()
        logger.error(f"Reset error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Запуск приложения
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=int(os.getenv("PORT", 8000)),
        reload=True
    )
