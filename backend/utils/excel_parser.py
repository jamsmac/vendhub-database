"""
Утилиты для парсинга Excel файлов
"""
import re
from datetime import datetime
from typing import List, Tuple, Optional, Dict, Any
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO


def parse_excel_file(file_content: bytes) -> Tuple[List[str], List[List[Any]]]:
    """
    Парсинг Excel файла

    Args:
        file_content: Байты файла

    Returns:
        Tuple[headers, rows]: Заголовки и строки данных
    """
    # Загрузка workbook
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    sheet = wb.active

    # Извлечение данных
    all_rows = list(sheet.iter_rows(values_only=True))

    if len(all_rows) < 2:
        raise ValueError("Файл должен содержать минимум заголовок и одну строку данных")

    # Заголовки (первая строка)
    headers = [str(cell) if cell is not None else f"Column_{i+1}"
               for i, cell in enumerate(all_rows[0])]

    # Данные (все строки кроме первой)
    rows = []
    for row in all_rows[1:]:
        # Пропускаем полностью пустые строки
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        # Конвертируем в строку и обрабатываем
        processed_row = [
            str(cell).strip() if cell is not None else ''
            for cell in row
        ]
        rows.append(processed_row)

    wb.close()
    return headers, rows


def extract_date_from_row(row: List[str]) -> Optional[datetime]:
    """
    Извлечение даты из строки данных

    Args:
        row: Строка данных

    Returns:
        datetime объект или None
    """
    # Форматы дат для проверки
    date_patterns = [
        (r'(\d{4})-(\d{1,2})-(\d{1,2})', '%Y-%m-%d'),      # YYYY-MM-DD
        (r'(\d{1,2})\.(\d{1,2})\.(\d{4})', '%d.%m.%Y'),    # DD.MM.YYYY
        (r'(\d{1,2})/(\d{1,2})/(\d{4})', '%m/%d/%Y'),      # MM/DD/YYYY
        (r'(\d{1,2})-(\d{1,2})-(\d{4})', '%d-%m-%Y'),      # DD-MM-YYYY
    ]

    # Проверяем каждую ячейку в строке
    for cell in row:
        if not cell or not isinstance(cell, str):
            continue

        cell = cell.strip()

        # Пробуем распарсить по паттернам
        for pattern, date_format in date_patterns:
            match = re.search(pattern, cell)
            if match:
                try:
                    date_str = match.group(0)
                    parsed_date = datetime.strptime(date_str, date_format)

                    # Проверяем разумность даты (2000-2100)
                    if 2000 <= parsed_date.year <= 2100:
                        return parsed_date
                except (ValueError, AttributeError):
                    continue

    return None


def get_period_from_date(date: datetime) -> str:
    """
    Получить период (YYYY-MM) из даты

    Args:
        date: Datetime объект

    Returns:
        Строка в формате YYYY-MM
    """
    return date.strftime("%Y-%m")


def create_unique_key(row: List[str]) -> str:
    """
    Создать уникальный ключ для строки (для дедупликации)

    Args:
        row: Строка данных

    Returns:
        Хеш строка
    """
    import hashlib
    import json

    # Нормализуем данные
    normalized = [str(cell).strip().lower() for cell in row]

    # Создаем JSON и хешируем
    row_json = json.dumps(normalized, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(row_json.encode()).hexdigest()


def validate_excel_structure(headers: List[str], rows: List[List[Any]]) -> Dict[str, Any]:
    """
    Валидация структуры Excel файла

    Returns:
        Статистика по файлу
    """
    stats = {
        "total_columns": len(headers),
        "total_rows": len(rows),
        "empty_cells": 0,
        "has_dates": False,
        "column_types": {}
    }

    # Подсчет статистики
    for row in rows:
        for i, cell in enumerate(row):
            if not cell or cell == '':
                stats["empty_cells"] += 1

            # Определяем тип данных колонки
            col_name = headers[i] if i < len(headers) else f"Column_{i+1}"
            if col_name not in stats["column_types"]:
                stats["column_types"][col_name] = {"numeric": 0, "text": 0, "date": 0}

            # Пробуем определить тип
            try:
                float(cell)
                stats["column_types"][col_name]["numeric"] += 1
            except (ValueError, TypeError):
                if extract_date_from_row([cell]):
                    stats["column_types"][col_name]["date"] += 1
                    stats["has_dates"] = True
                else:
                    stats["column_types"][col_name]["text"] += 1

    return stats
